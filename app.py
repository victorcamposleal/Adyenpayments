from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
import os
import hmac
import hashlib
import base64
import logging

app = Flask(__name__)
excel_file = "adyen_webhooks.xlsx"
HMAC_KEY = os.getenv("ADYEN_HMAC_KEY")

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Crear archivo Excel si no existe
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Webhooks"
    ws.append(["EventCode", "Success", "PSPReference", "MerchantAccount", "Currency", "Amount"])
    wb.save(excel_file)

def is_valid_hmac(notification, hmac_key):
    try:
        data_to_sign = ":".join([
            notification.get("pspReference", ""),
            notification.get("originalReference", ""),
            notification.get("merchantAccountCode", ""),
            notification.get("merchantReference", ""),
            str(notification.get("amount", {}).get("value", "")),
            notification.get("amount", {}).get("currency", ""),
            notification.get("eventCode", ""),
            notification.get("success", "")
        ])
        hmac_calculated = base64.b64encode(
            hmac.new(base64.b64decode(hmac_key), data_to_sign.encode("utf-8"), hashlib.sha256).digest()
        ).decode()
        hmac_received = notification.get("additionalData", {}).get("hmacSignature")
        return hmac_calculated == hmac_received
    except Exception as e:
        logging.error(f"Error verificando HMAC: {e}")
        return False
@app.route('/adyen-webhook', methods=['POST'])
def adyen_webhook():
    try:
        data = request.get_json()
        logging.info(f"Webhook recibido: {data}")

        if "notificationItems" in data:
            wb = load_workbook(excel_file)
            ws = wb["Webhooks"]

            for item in data["notificationItems"]:
                notif = item["NotificationRequestItem"]

                if not is_valid_hmac(notif, HMAC_KEY):
                    logging.warning("Firma HMAC inválida. Notificación ignorada.")
                    continue

                ws.append([
                    notif.get("eventCode"),
                    notif.get("success"),
                    notif.get("pspReference"),
                    notif.get("merchantAccountCode"),
                    notif.get("amount", {}).get("currency"),
                    notif.get("amount", {}).get("value")
                ])

            wb.save(excel_file)

        return jsonify({"status": "[accepted]"}), 200

    except Exception as e:
        logging.error(f"Error procesando webhook: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)